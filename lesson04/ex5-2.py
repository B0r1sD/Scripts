#!/usr/bin/python

import os
import openpyxl
import humanfriendly
import re
import fileinput

#initialise
wb = openpyxl.Workbook()

#activate worksheet
ws = wb.active

#add header row
ws.append(["sample_ID","filesize"])

#scan folder

'''inputfile = fileinput.input("exercise-5-2-htseqcount/all_counts.txt")
for line in inputfile:
    print(line)'''



location = "/home/guest/BITsemA/Scripting/lesson04"
os.chdir(location)

print("current working directory: {}".format(os.getcwd()))

'''walk = os.walk(location + "/exercise-5-2-htseqcount")
for x in walk:
    print(x)
    for y in x:
        print(y)'''

list_dir = os.listdir(location + "/exercise-5-2-htseqcount")
#print(list_dir)

listID = []
sizeID = []
i = 1
a = 0

for x in list_dir:
    size = os.stat("exercise-5-2-htseqcount/"+ x)
    realsize = size.st_size
    #print(realsize)
    hsize = humanfriendly.format_size(realsize)
    #ws.append([x])
    splitted = x.split('_')
    #print(splitted)
    if re.search('^count', splitted[0]):
        i += 1
        listID.append(splitted[1])
        sizeID.append(hsize)
        #print("sample_ID: {}".format(splitted[1]))
        #print("sample size: {}".format(hsize))
        '''ws['A' + str(i)] = listID[int(a)]
        ws['B' + str(i)] = sizeID[int(a)]
        a += 1'''
        ws.append([splitted[1],hsize])

print(listID)
print(sizeID)
#save excel file
wb.save("sample_ids_sizes_V1.xlsx")